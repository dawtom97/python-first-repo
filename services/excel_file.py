import os
from openpyxl import Workbook, load_workbook


def save_to_excel(filename, data):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(list(data.keys()))

    ws.append(list(data.values()))
    wb.save(filename)


# def create_excel():
#     values = [23,21,22,21,19,20,23,40,17]
#
#     wb = Workbook()
#     sheet = wb.active
#     sheet.title = "Temperatury"
#
#     sheet.append(["Pomiar nr.", "Temperatura", "Test"])
#
#     for index, value in enumerate(values,start=1):
#         sheet.append([index, value, f"test{index}"])
#     wb.save("testowy.xlsx")
# create_excel()
